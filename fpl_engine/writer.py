"""
Excel writer: builds FPL_Master.xlsx with all tabs, colour heatmaps and
conditional formatting, using openpyxl directly for full control.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule

from . import config as C

# ---- palette ----
NAVY = "1F2A44"
HEADER_FILL = PatternFill("solid", fgColor=NAVY)
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
TITLE_FONT = Font(bold=True, size=14, color=NAVY)
SUB_FONT = Font(bold=True, size=11, color=NAVY)
NOTE_FONT = Font(italic=True, size=9, color="666666")
BAND = PatternFill("solid", fgColor="F2F5FA")
THIN = Side(style="thin", color="D9DEE8")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

FDR_FILLS = {
    1: PatternFill("solid", fgColor="1A9850"),   # dark green - easiest
    2: PatternFill("solid", fgColor="91CF60"),
    3: PatternFill("solid", fgColor="FFFFBF"),   # neutral
    4: PatternFill("solid", fgColor="FC8D59"),
    5: PatternFill("solid", fgColor="D73027"),   # red - hardest
}
FDR_FONT = {1: Font(color="FFFFFF", size=9, bold=True),
            2: Font(color="0B3D0B", size=9),
            3: Font(color="333333", size=9),
            4: Font(color="3D1400", size=9),
            5: Font(color="FFFFFF", size=9, bold=True)}


def _auto_width(ws, df_cols, start_col=1, max_w=32):
    for j, col in enumerate(df_cols):
        letter = get_column_letter(start_col + j)
        w = max(len(str(col)) + 2, 8)
        ws.column_dimensions[letter].width = min(w, max_w)


def write_table(ws, df, start_row=1, start_col=1, title=None,
                autofilter=False, freeze=False, band=True):
    """Write a DataFrame as a styled table; return the row after the table."""
    r = start_row
    if title:
        cell = ws.cell(row=r, column=start_col, value=title)
        cell.font = SUB_FONT
        r += 1
    header_row = r
    for j, col in enumerate(df.columns):
        cell = ws.cell(row=r, column=start_col + j, value=str(col))
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
        cell.border = BORDER
    r += 1
    for _, row in df.iterrows():
        for j, col in enumerate(df.columns):
            v = row[col]
            cell = ws.cell(row=r, column=start_col + j, value=v)
            cell.border = BORDER
            cell.font = Font(size=9)
            if band and (r - header_row) % 2 == 0:
                cell.fill = BAND
        r += 1
    if autofilter:
        first = get_column_letter(start_col)
        last = get_column_letter(start_col + len(df.columns) - 1)
        ws.auto_filter.ref = f"{first}{header_row}:{last}{r - 1}"
    if freeze:
        ws.freeze_panes = ws.cell(row=header_row + 1, column=start_col + 3)
    return r, header_row


def _colscale_green_high(ws, col_letter, r0, r1):
    ws.conditional_formatting.add(
        f"{col_letter}{r0}:{col_letter}{r1}",
        ColorScaleRule(start_type="min", start_color="F8696B",
                       mid_type="percentile", mid_value=50, mid_color="FFEB84",
                       end_type="max", end_color="63BE7B"))


def _databar(ws, col_letter, r0, r1, color="5B9BD5"):
    ws.conditional_formatting.add(
        f"{col_letter}{r0}:{col_letter}{r1}",
        DataBarRule(start_type="min", end_type="max", color=color))


# ---------------------------------------------------------------------------
def build_workbook(ctx):
    """
    ctx keys: df, teams_df, fixture_matrix (dict), tracker_events, rankings,
    optimiser, my_squad_df, meta
    """
    wb = Workbook()
    wb.remove(wb.active)

    _sheet_dashboard(wb, ctx)
    _sheet_player_db(wb, ctx)
    _sheet_teams(wb, ctx)
    _sheet_fixture_tracker(wb, ctx)
    _sheet_watchlist(wb, ctx)
    _sheet_rankings(wb, ctx)
    _sheet_optimizer(wb, ctx)
    _sheet_setpieces(wb, ctx)
    _sheet_ownership(wb, ctx)
    _sheet_chips(wb, ctx)
    _sheet_my_squad(wb, ctx)

    try:
        wb.save(C.OUTPUT_XLSX)
        return C.OUTPUT_XLSX
    except PermissionError:
        alt = C.OUTPUT_XLSX.with_name("FPL_Master_new.xlsx")
        wb.save(alt)
        print(f"  ! {C.OUTPUT_XLSX.name} was locked (open in Excel?) — "
              f"wrote {alt.name} instead")
        return alt


# ---- individual sheets ----------------------------------------------------
def _sheet_dashboard(wb, ctx):
    ws = wb.create_sheet("Dashboard")
    ws.sheet_view.showGridLines = False
    m = ctx["meta"]
    ws["A1"] = "FPL COMMAND CENTRE"
    ws["A1"].font = Font(bold=True, size=18, color=NAVY)
    ws["A2"] = (f"Season {m['season']}  |  Next: GW{m['next_event']}  |  "
                f"GWs played: {m['games_played']}  |  Updated: {m['updated']}")
    ws["A2"].font = NOTE_FONT
    ws["A3"] = ("xP = model's Expected Points. xP_next = coming GW, "
                "xP_H = next %d GWs. Green = better." % C.HORIZON)
    ws["A3"].font = NOTE_FONT

    r = 5
    df = ctx["df"]
    # Top 10 overall by xP_next
    top = (df[df["p_start%"] >= 40]
           .sort_values("xP_next", ascending=False)
           .head(12)[["name", "team", "pos", "price", "own%",
                      "xP_next", "xP_H", "form"]])
    rr, hr = write_table(ws, top, r, 1, "Top picks — next GW (xP)")
    _colscale_green_high(ws, get_column_letter(6), hr + 1, rr - 1)
    _colscale_green_high(ws, get_column_letter(7), hr + 1, rr - 1)

    # Captain picks
    cap = (df[df["p_start%"] >= 60]
           .sort_values("xP_next", ascending=False)
           .head(6)[["name", "team", "xP_next", "own%"]])
    rr2, hr2 = write_table(ws, cap, r, 10, "Captain candidates (next GW)")

    r = rr + 2
    # Best value & best differential side by side
    val = (df[(df["p_start%"] >= 50) & (df["price"] <= 8)]
           .sort_values("value_H", ascending=False)
           .head(10)[["name", "team", "pos", "price", "xP_H", "value_H"]])
    rr3, hr3 = write_table(ws, val, r, 1, "Best value (xP per £m, ≤£8.0)")
    _databar(ws, get_column_letter(6), hr3 + 1, rr3 - 1)

    dif = (df[(df["own%"] <= C.DIFFERENTIAL_MAX_OWNERSHIP) &
              (df["p_start%"] >= 55)]
           .sort_values("xP_H", ascending=False)
           .head(10)[["name", "team", "pos", "own%", "xP_H"]])
    write_table(ws, dif, r, 9, f"Best differentials (≤{int(C.DIFFERENTIAL_MAX_OWNERSHIP)}% owned)")

    r = rr3 + 2
    # Best DEFCON set-and-forget
    dc = (df[(df["pos"].isin(["DEF", "MID"])) & (df["p_start%"] >= 55)]
          .sort_values("xP_defcon", ascending=False)
          .head(10)[["name", "team", "pos", "price", "DEFCON90",
                     "xP_defcon", "xP_H"]])
    write_table(ws, dc, r, 1, "Best Defensive-Contribution (DEFCON) picks")

    _auto_width(ws, ["name123456", "team", "pos", "price", "own%",
                     "xP_next", "xP_H", "form"], 1)
    for col in "JKLM":
        ws.column_dimensions[col].width = 12
    ws.column_dimensions["A"].width = 20


def _sheet_player_db(wb, ctx):
    ws = wb.create_sheet("Player DB")
    ws.sheet_view.showGridLines = False
    cols = ["name", "team", "pos", "price", "own%", "form", "pts", "ppg",
            "nailed", "p_start%", "xG90", "xA90", "xGI90", "DEFCON90",
            "saves90", "bonus90", "pen", "fk", "ck",
            "xP_next", "xP_H", "xP_attack", "xP_cs", "xP_defcon", "xP_bonus",
            "value_next", "value_H", "status", "news"]
    df = ctx["df"][cols].sort_values("xP_H", ascending=False)
    rr, hr = write_table(ws, df, 1, 1, None, autofilter=True, freeze=True,
                         band=False)
    for name in ("xP_next", "xP_H", "xP_attack", "xP_cs", "xP_defcon"):
        c = get_column_letter(cols.index(name) + 1)
        _colscale_green_high(ws, c, hr + 1, rr - 1)
    _auto_width(ws, cols, 1)
    ws.column_dimensions[get_column_letter(cols.index("news") + 1)].width = 40
    ws.column_dimensions["A"].width = 18


def _sheet_teams(wb, ctx):
    ws = wb.create_sheet("Teams")
    ws.sheet_view.showGridLines = False
    df = ctx["teams_df"]
    rr, hr = write_table(ws, df, 1, 1, "Team ratings & fixture outlook",
                         autofilter=True)
    # colour attack (high good) and defence-conceded (low good), and Next6 FDR
    if "AttackRating" in df.columns:
        _colscale_green_high(ws, get_column_letter(df.columns.get_loc("AttackRating") + 1), hr + 1, rr - 1)
    if "Next6_FDR" in df.columns:
        col = get_column_letter(df.columns.get_loc("Next6_FDR") + 1)
        ws.conditional_formatting.add(
            f"{col}{hr+1}:{col}{rr-1}",
            ColorScaleRule(start_type="num", start_value=1, start_color="1A9850",
                           mid_type="num", mid_value=3, mid_color="FFFFBF",
                           end_type="num", end_value=5, end_color="D73027"))
    _auto_width(ws, list(df.columns), 1)
    ws.column_dimensions["A"].width = 18


def _sheet_fixture_tracker(wb, ctx):
    ws = wb.create_sheet("Fixture Tracker")
    ws.sheet_view.showGridLines = False
    ws["A1"] = "FIXTURE TRACKER — green = easy, red = hard (DBL = double GW, blank = no fixture)"
    ws["A1"].font = SUB_FONT
    matrix = ctx["fixture_matrix"]      # {team_short: {event: {'text','fdr'}}}
    events = ctx["tracker_evlist"]
    start = 3
    ws.cell(row=start, column=1, value="Team").fill = HEADER_FILL
    ws.cell(row=start, column=1).font = HEADER_FONT
    for j, ev in enumerate(events):
        c = ws.cell(row=start, column=2 + j, value=f"GW{ev}")
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(2 + j)].width = 11
    r = start + 1
    for team in sorted(matrix.keys()):
        tc = ws.cell(row=r, column=1, value=team)
        tc.font = Font(bold=True, size=9)
        tc.border = BORDER
        for j, ev in enumerate(events):
            cell = ws.cell(row=r, column=2 + j)
            info = matrix[team].get(ev)
            if not info:
                cell.value = "-"
                cell.fill = PatternFill("solid", fgColor="EDEDED")
                cell.font = Font(color="AAAAAA", size=9)
            else:
                cell.value = info["text"]
                fdr = info["fdr"]
                cell.fill = FDR_FILLS[fdr]
                cell.font = FDR_FONT[fdr]
            cell.alignment = Alignment(horizontal="center")
            cell.border = BORDER
        r += 1
    ws.column_dimensions["A"].width = 8
    ws.freeze_panes = ws.cell(row=start + 1, column=2)


def _sheet_watchlist(wb, ctx):
    ws = wb.create_sheet("Watchlist")
    ws.sheet_view.showGridLines = False
    ws["A1"] = "WATCHLIST — best targets in every club (ranked by xP over horizon)"
    ws["A1"].font = SUB_FONT
    df = ctx["watchlist_df"]
    rr, hr = write_table(ws, df, 3, 1, None, autofilter=True)
    _colscale_green_high(ws, get_column_letter(df.columns.get_loc("xP_H") + 1),
                         hr + 1, rr - 1)
    _auto_width(ws, list(df.columns), 1)
    ws.column_dimensions["B"].width = 18


def _sheet_rankings(wb, ctx):
    ws = wb.create_sheet("Rankings")
    ws.sheet_view.showGridLines = False
    ws["A1"] = "RANKINGS — best options by position & route"
    ws["A1"].font = SUB_FONT
    r = 3
    col = 1
    for title, rdf in ctx["rankings"]:
        rr, hr = write_table(ws, rdf, r, col, title)
        if "xP_H" in rdf.columns:
            _colscale_green_high(
                ws, get_column_letter(col + rdf.columns.get_loc("xP_H")),
                hr + 1, rr - 1)
        # lay out in two columns of tables
        if col == 1:
            col = 9
        else:
            col = 1
            r = rr + 2
        if col == 9:
            # keep same r for the right-hand table
            pass
    for i in range(1, 16):
        ws.column_dimensions[get_column_letter(i)].width = 12
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["I"].width = 18


def _sheet_optimizer(wb, ctx):
    ws = wb.create_sheet("Optimizer")
    ws.sheet_view.showGridLines = False
    opt = ctx["optimiser"]
    ws["A1"] = "OPTIMISED SQUAD — best legal 15 under £100m (max 3/club)"
    ws["A1"].font = SUB_FONT
    if opt is None:
        ws["A3"] = "Optimiser could not find a solution."
        return
    ws["A2"] = (f"Total cost £{opt['cost']:.1f}m  |  XI xP_H {opt['xi_xp']:.1f}  |  "
                f"Captain: {opt['captain']}  |  status: {opt['status']}")
    ws["A2"].font = NOTE_FONT
    rr, hr = write_table(ws, opt["df"], 4, 1, None)
    _colscale_green_high(ws, get_column_letter(opt["df"].columns.get_loc("xP_H") + 1),
                         hr + 1, rr - 1)
    _auto_width(ws, list(opt["df"].columns), 1)
    ws.column_dimensions["A"].width = 18


def _sheet_setpieces(wb, ctx):
    ws = wb.create_sheet("Set Pieces")
    ws.sheet_view.showGridLines = False
    ws["A1"] = "SET-PIECE & PENALTY TAKERS — order 1 = first choice (hidden xG source)"
    ws["A1"].font = SUB_FONT
    df = ctx["setpiece_df"]
    write_table(ws, df, 3, 1, None, autofilter=True)
    _auto_width(ws, list(df.columns), 1)
    ws.column_dimensions["B"].width = 18


def _sheet_ownership(wb, ctx):
    ws = wb.create_sheet("Ownership")
    ws.sheet_view.showGridLines = False
    ws["A1"] = "OWNERSHIP — template (high) vs differentials (low). Cover template, attack with diffs."
    ws["A1"].font = SUB_FONT
    df = ctx["df"].sort_values("own%", ascending=False).head(60)[
        ["name", "team", "pos", "price", "own%", "xP_H", "xP_next", "form"]]
    rr, hr = write_table(ws, df, 3, 1, None, autofilter=True)
    _databar(ws, get_column_letter(5), hr + 1, rr - 1, color="9E480E")
    _colscale_green_high(ws, get_column_letter(6), hr + 1, rr - 1)
    _auto_width(ws, list(df.columns), 1)
    ws.column_dimensions["A"].width = 18


def _sheet_chips(wb, ctx):
    ws = wb.create_sheet("Chip Planner")
    ws.sheet_view.showGridLines = False
    ws["A1"] = "CHIP PLANNER — map chips to double (DBL) & blank (BLK) gameweeks"
    ws["A1"].font = SUB_FONT
    write_table(ws, ctx["chips_df"], 3, 1, None)
    notes = [
        "",
        "STRATEGY NOTES",
        "• Bench Boost: play in a double GW where all 15 have 2 fixtures & good FDR.",
        "• Triple Captain: a premium (e.g. Haaland/Salah-type) at home in a double GW.",
        "• Free Hit: best in a big blank GW to field a full XI, or to attack a double.",
        "• Wildcard: use when 4+ of your team need changing, ideally before a good fixture swing.",
        "• Watch the Fixture Tracker: buy a team BEFORE a green run, sell before a red run.",
    ]
    rbase = 3 + len(ctx["chips_df"]) + 3
    for i, n in enumerate(notes):
        cell = ws.cell(row=rbase + i, column=1, value=n)
        cell.font = SUB_FONT if n == "STRATEGY NOTES" else NOTE_FONT
    for i in range(1, 6):
        ws.column_dimensions[get_column_letter(i)].width = 16
    ws.column_dimensions["A"].width = 70


def _sheet_my_squad(wb, ctx):
    ws = wb.create_sheet("My Squad")
    ws.sheet_view.showGridLines = False
    ws["A1"] = "MY SQUAD — your team scored by the model"
    ws["A1"].font = Font(bold=True, size=14, color=NAVY)
    df = ctx["my_squad_df"]
    rr, hr = write_table(ws, df, 3, 1, None)
    if "xP_H" in df.columns:
        _colscale_green_high(ws, get_column_letter(df.columns.get_loc("xP_H") + 1),
                             hr + 1, rr - 1)
    r = rr + 2
    for line in ctx["my_squad_notes"]:
        cell = ws.cell(row=r, column=1, value=line)
        cell.font = SUB_FONT if line.startswith(("VERDICT", "SUGGESTIONS", "FLAGS")) else NOTE_FONT
        r += 1
    _auto_width(ws, list(df.columns), 1)
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions[get_column_letter(len(df.columns))].width = 40
